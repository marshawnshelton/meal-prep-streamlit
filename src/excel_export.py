#!/usr/bin/env python3
"""
Excel Export Module
Exports shopping lists to Excel with separate sheets per store
"""

import yaml
from pathlib import Path
from datetime import datetime
from typing import Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl --break-system-packages")


class ExcelExporter:
    def __init__(self):
        """Initialize Excel exporter"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl --break-system-packages")
        
        # Define styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.store_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.store_header_font = Font(color="FFFFFF", bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _create_summary_sheet(self, wb: Workbook, shopping_list: Dict) -> None:
        """Create overview summary sheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Title
        ws['A1'] = "Bi-Weekly Shopping List"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Key info
        row = 3
        info = [
            ("Planning Period:", shopping_list['meal_plan_dates']),
            ("People:", shopping_list['people']),
            ("Budget:", f"${shopping_list['budget']}"),
            ("Stores:", len(shopping_list['stores']))
        ]
        
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Store breakdown
        row += 2
        ws[f'A{row}'] = "Stores to Visit:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Store"
        ws[f'B{row}'] = "Type"
        ws[f'C{row}'] = "Items"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        for store_name, store_data in shopping_list['stores'].items():
            ws[f'A{row}'] = store_name.replace('_', ' ').title()
            ws[f'B{row}'] = store_data['store_info'].get('type', 'N/A').title()
            ws[f'C{row}'] = len(store_data['items'])
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_store_sheet(self, wb: Workbook, store_name: str, store_data: Dict) -> None:
        """Create individual sheet for each store"""
        # Clean store name for sheet title
        sheet_title = store_name.replace('_', ' ').title()[:31]  # Excel limit
        ws = wb.create_sheet(title=sheet_title)
        
        # Store header
        ws['A1'] = sheet_title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.store_header_fill
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Store info
        ws['A2'] = f"Type: {store_data['store_info'].get('type', 'N/A').title()}"
        ws['A2'].font = Font(italic=True)
        
        # Column headers
        row = 4
        headers = ['Item', 'Amount', 'Unit', 'Used In (Recipes)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Items
        row += 1
        for item in store_data['items']:
            ws.cell(row=row, column=1, value=item['item'].title())
            ws.cell(row=row, column=2, value=str(item['amount']))
            ws.cell(row=row, column=3, value=item['unit'])
            
            # Format recipes list
            recipes = item.get('used_in', [])
            recipes_str = ', '.join(recipes[:2])  # First 2 recipes
            if len(recipes) > 2:
                recipes_str += f" +{len(recipes)-2} more"
            ws.cell(row=row, column=4, value=recipes_str)
            
            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        
        # Add shopping checklist column
        ws.cell(row=4, column=5, value='✓ Got It')
        ws.cell(row=4, column=5).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws.cell(row=4, column=5).font = Font(bold=True)
        ws.cell(row=4, column=5).border = self.border
        ws.cell(row=4, column=5).alignment = Alignment(horizontal='center')
        ws.column_dimensions['E'].width = 12
    
    def _create_master_list_sheet(self, wb: Workbook, shopping_list: Dict) -> None:
        """Create a master list combining all stores"""
        ws = wb.create_sheet(title="Master List")
        
        # Title
        ws['A1'] = "Master Shopping List - All Stores"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        row = 3
        headers = ['Store', 'Item', 'Amount', 'Unit', 'Used In']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        # Aggregate all items
        row += 1
        for store_name, store_data in shopping_list['stores'].items():
            for item in store_data['items']:
                ws.cell(row=row, column=1, value=store_name.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=item['item'].title())
                ws.cell(row=row, column=3, value=str(item['amount']))
                ws.cell(row=row, column=4, value=item['unit'])
                
                recipes = item.get('used_in', [])
                recipes_str = ', '.join(recipes[:2])
                if len(recipes) > 2:
                    recipes_str += f" +{len(recipes)-2} more"
                ws.cell(row=row, column=5, value=recipes_str)
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35
    
    def export_to_excel(self, shopping_list: Dict, output_dir: str = "output") -> str:
        """Export shopping list to Excel file"""
        Path(output_dir).mkdir(exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb, shopping_list)
        
        # Create individual store sheets
        for store_name, store_data in shopping_list['stores'].items():
            self._create_store_sheet(wb, store_name, store_data)
        
        # Create master list
        self._create_master_list_sheet(wb, shopping_list)
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{output_dir}/shopping_list_{timestamp}.xlsx"
        wb.save(filename)
        
        return filename


if __name__ == "__main__":
    # Test the Excel exporter
    import sys
    sys.path.append('.')
    from src.meal_planner import MealPlanner
    from src.shopping_list import ShoppingListGenerator
    
    # Generate meal plan and shopping list
    planner = MealPlanner()
    meal_plan = planner.generate_meal_plan()
    
    generator = ShoppingListGenerator()
    shopping_list = generator.generate_shopping_list(meal_plan)
    
    # Export to Excel
    exporter = ExcelExporter()
    excel_file = exporter.export_to_excel(shopping_list)
    
    print(f"Excel shopping list created: {excel_file}")
    print("\nThe file includes:")
    print("  - Summary sheet with overview")
    print("  - Individual sheets for each store")
    print("  - Master list combining all stores")
    print("  - Checklist column for marking items as purchased")
